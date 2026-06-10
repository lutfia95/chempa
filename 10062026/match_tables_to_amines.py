#!/usr/bin/env python3
"""Match workbook tables S1/S2/S3 to primary, secondary, and aromatic amines."""

from __future__ import annotations

import argparse
import csv
import sys
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit("Missing dependency: openpyxl.") from exc

try:
    from rdkit import Chem
except ImportError as exc:
    raise SystemExit(
        "Missing dependency: RDKit. Run with the chempa-convert Conda environment."
    ) from exc


DEFAULT_WORKBOOK = "Science 2024, 384, eadk5864_tables_s1_to_s7.xlsx"


@dataclass(frozen=True)
class AmineConfig:
    name: str
    filename: str


AMINE_CONFIGS = (
    AmineConfig("primary", "primary_aliphatic_amines.tsv"),
    AmineConfig("secondary", "secondary_aliphatic_amines.tsv"),
    AmineConfig("aromatic", "aromatic_amines.tsv"),
    AmineConfig("aromatic", "others.tsv"),
)

OUTPUT_COLUMNS = [
    "table",
    "source_excel_row",
    "fragId",
    "catalog_id",
    "accession",
    "geneName",
    "source_smiles_column",
    "source_smiles",
    "canonical_smiles",
    "matched_amine_SMILES",
    "amine_Label",
    "amine_File",
    "amine_Index",
]


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def canonical_smiles(smiles: str) -> str:
    molecule = Chem.MolFromSmiles(smiles) if smiles else None
    if molecule is None:
        return ""
    return Chem.MolToSmiles(molecule, canonical=True, isomericSmiles=True)


def read_amine_index(path: Path) -> tuple[dict[str, list[dict[str, str]]], int]:
    index: dict[str, list[dict[str, str]]] = defaultdict(list)
    invalid = 0

    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        if "SMILES" not in (reader.fieldnames or []):
            raise ValueError(f"Missing SMILES column in {path}")

        for row_number, row in enumerate(reader, start=2):
            smiles = clean(row.get("SMILES"))
            canonical = canonical_smiles(smiles)
            if not canonical:
                invalid += 1
                continue
            record = dict(row)
            record["_tsv_row"] = str(row_number)
            index[canonical].append(record)

    return dict(index), invalid


def worksheet_records(workbook: Any, sheet_name: str) -> list[dict[str, Any]]:
    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(rows)]
    records: list[dict[str, Any]] = []
    for excel_row, values in enumerate(rows, start=2):
        padded = tuple(values) + (None,) * (len(headers) - len(values))
        record = dict(zip(headers, padded))
        record["_excel_row"] = excel_row
        records.append(record)
    return records


def source_records(workbook: Any, table_name: str) -> Iterable[dict[str, str]]:
    if table_name == "table S1":
        for row in worksheet_records(workbook, table_name):
            yield {
                "table": table_name,
                "source_excel_row": str(row["_excel_row"]),
                "fragId": clean(row.get("fragId")),
                "catalog_id": "",
                "accession": "",
                "geneName": "",
                "source_smiles_column": "SMILES",
                "source_smiles": clean(row.get("SMILES")),
            }
        return

    if table_name == "table S3":
        for row in worksheet_records(workbook, table_name):
            yield {
                "table": table_name,
                "source_excel_row": str(row["_excel_row"]),
                "fragId": clean(row.get("fid")),
                "catalog_id": clean(row.get("catalog_id")),
                "accession": "",
                "geneName": "",
                "source_smiles_column": "smiles",
                "source_smiles": clean(row.get("smiles")),
            }
        return

    if table_name == "table S2":
        s1_smiles_by_frag_id = {
            clean(row.get("fragId")): clean(row.get("SMILES"))
            for row in worksheet_records(workbook, "table S1")
        }
        for row in worksheet_records(workbook, table_name):
            frag_id = clean(row.get("fragId"))
            yield {
                "table": table_name,
                "source_excel_row": str(row["_excel_row"]),
                "fragId": frag_id,
                "catalog_id": "",
                "accession": clean(row.get("accession")),
                "geneName": clean(row.get("geneName")),
                "source_smiles_column": "table S1.SMILES via fragId",
                "source_smiles": s1_smiles_by_frag_id.get(frag_id, ""),
            }
        return

    raise ValueError(f"Unsupported table: {table_name}")


def write_matches(
    path: Path,
    sources: Iterable[dict[str, str]],
    amine_index: dict[str, list[dict[str, str]]],
) -> tuple[int, int, int]:
    path.parent.mkdir(parents=True, exist_ok=True)
    source_count = 0
    matched_source_count = 0
    output_count = 0

    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=OUTPUT_COLUMNS,
            delimiter="\t",
            lineterminator="\n",
        )
        writer.writeheader()

        for source in sources:
            source_count += 1
            canonical = canonical_smiles(source["source_smiles"])
            matches = amine_index.get(canonical, []) if canonical else []
            if matches:
                matched_source_count += 1

            for amine in matches:
                output = dict(source)
                output.update(
                    {
                        "canonical_smiles": canonical,
                        "matched_amine_SMILES": clean(amine.get("SMILES")),
                        "amine_Label": clean(amine.get("Label")),
                        "amine_File": clean(amine.get("File")),
                        "amine_Index": clean(amine.get("Index")),
                    }
                )
                writer.writerow(output)
                output_count += 1

    return source_count, matched_source_count, output_count


def parse_args() -> argparse.Namespace:
    script_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(
        description=(
            "Exact canonical-SMILES matching of workbook tables S1/S2/S3 against "
            "primary, secondary, and aromatic amine TSV files."
        )
    )
    parser.add_argument(
        "workbook",
        nargs="?",
        type=Path,
        default=script_dir / DEFAULT_WORKBOOK,
    )
    parser.add_argument(
        "--amines-dir",
        type=Path,
        default=script_dir / "amines",
        help="Directory containing the three source amine TSV files",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=script_dir / "amines",
        help="Directory for the nine matched TSV files",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    workbook_path = args.workbook.expanduser().resolve()
    amines_dir = args.amines_dir.expanduser().resolve()
    output_dir = args.output_dir.expanduser().resolve()

    if not workbook_path.is_file():
        print(f"Workbook not found: {workbook_path}", file=sys.stderr)
        return 2

    amine_indexes: dict[str, dict[str, list[dict[str, str]]]] = {}
    for config in AMINE_CONFIGS:
        path = amines_dir / config.filename
        if not path.is_file():
            print(f"Amine TSV not found: {path}", file=sys.stderr)
            return 2
        index, invalid = read_amine_index(path)
        amine_indexes[config.name] = index
        molecule_count = sum(len(records) for records in index.values())
        print(
            f"[INFO] {config.name}: indexed {molecule_count} molecules "
            f"({len(index)} unique canonical SMILES, {invalid} invalid)"
        )

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        required_sheets = {"table S1", "table S2", "table S3"}
        missing_sheets = required_sheets.difference(workbook.sheetnames)
        if missing_sheets:
            raise ValueError(
                f"Workbook is missing sheets: {', '.join(sorted(missing_sheets))}"
            )

        for table_number, table_name in ((1, "table S1"), (2, "table S2"), (3, "table S3")):
            sources = list(source_records(workbook, table_name))
            for config in AMINE_CONFIGS:
                output_path = (
                    output_dir
                    / f"table_s{table_number}_matched_{config.name}.tsv"
                )
                source_count, matched_count, output_count = write_matches(
                    output_path,
                    sources,
                    amine_indexes[config.name],
                )
                print(
                    f"[OK] {output_path.name}: {matched_count}/{source_count} "
                    f"source rows matched; {output_count} output rows"
                )
    finally:
        workbook.close()

    print(f"[DONE] Wrote nine match files to {output_dir}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except ValueError as exc:
        print(f"[ERROR] {exc}", file=sys.stderr)
        raise SystemExit(1)
