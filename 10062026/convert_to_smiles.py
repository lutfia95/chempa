#!/usr/bin/env python3
"""Convert selected workbook SMILES columns to structure PDFs and TSV files."""

from __future__ import annotations

import argparse
import csv
import re
import sys
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable, Sequence

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit("Missing dependency: openpyxl. Install it with 'pip install openpyxl'.") from exc

try:
    from rdkit import Chem
    from rdkit.Chem import AllChem
    from rdkit.Chem.Draw import rdMolDraw2D
except ImportError as exc:
    raise SystemExit("Missing dependency: RDKit. Install it with 'conda install -c conda-forge rdkit'.") from exc

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.utils import ImageReader
    from reportlab.pdfbase.pdfmetrics import stringWidth
    from reportlab.pdfgen import canvas
except ImportError as exc:
    raise SystemExit("Missing dependency: reportlab. Install it with 'pip install reportlab'.") from exc


DEFAULT_WORKBOOK = "Science 2024, 384, eadk5864_tables_s1_to_s7.xlsx"


@dataclass(frozen=True)
class SheetConfig:
    name: str
    smiles_column: str
    title_columns: tuple[str, ...]
    pdf_columns: tuple[str, ...] | None = None


SHEET_CONFIGS = (
    SheetConfig(
        name="table S1",
        smiles_column="SMILES Naked",
        title_columns=("fragId",),
    ),
    SheetConfig(
        name="table S3",
        smiles_column="smiles",
        title_columns=("catalog_id", "fid"),
        pdf_columns=(
            "catalog_id",
            "fid",
            "hasTertiaryAmine",
            "nHBD",
            "nHBA",
            "cLogP",
            "nHeteroAtoms",
            "RingCount",
            "nRotatableBonds",
            "nAromaticBonds",
            "nAcidicGroup",
            "nBasicGroup",
            "AtomicPolarizability",
            "MolWt",
            "TPSA",
            "MaxTrainSimilarity",
            "FpClusterId",
            "PhysChemClusterId",
        ),
    ),
    SheetConfig(
        name="table S4",
        smiles_column="smiles",
        title_columns=("fid",),
    ),
)


@dataclass
class MoleculeRow:
    excel_row: int
    source_smiles: str
    canonical_smiles: str
    mol: Chem.Mol | None
    error: str
    values: dict[str, Any]


def clean_text(value: Any) -> str:
    """Convert an Excel value to one TSV/PDF-safe line of text."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, float):
        return format(value, ".15g")
    return re.sub(r"\s+", " ", str(value)).strip()


def safe_stem(sheet_name: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "_", sheet_name).strip("_").lower()


def parse_smiles(value: Any) -> tuple[str, Chem.Mol | None, str]:
    source = clean_text(value)
    if not source:
        return "", None, "empty_smiles"

    mol = Chem.MolFromSmiles(source)
    if mol is None:
        return "", None, "rdkit_parse_failed"

    try:
        canonical = Chem.MolToSmiles(mol, canonical=True, isomericSmiles=True)
    except Exception as exc:
        return "", None, f"canonicalization_failed:{type(exc).__name__}"
    return canonical, mol, ""


def read_sheet_rows(
    workbook: Any,
    config: SheetConfig,
    limit: int | None = None,
) -> tuple[list[str], list[MoleculeRow]]:
    if config.name not in workbook.sheetnames:
        raise ValueError(f"Worksheet not found: {config.name}")

    worksheet = workbook[config.name]
    raw_headers = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [clean_text(value) for value in raw_headers]
    if not all(headers):
        raise ValueError(f"{config.name} contains one or more empty header cells")
    if len(set(headers)) != len(headers):
        raise ValueError(f"{config.name} contains duplicate column names")
    if config.smiles_column not in headers:
        raise ValueError(
            f"Column {config.smiles_column!r} not found in {config.name}; "
            f"available columns: {', '.join(headers)}"
        )

    smiles_index = headers.index(config.smiles_column)
    parsed_rows: list[MoleculeRow] = []
    rows = worksheet.iter_rows(min_row=2, values_only=True)
    for output_index, values in enumerate(rows, start=1):
        if limit is not None and output_index > limit:
            break
        padded_values = tuple(values) + (None,) * (len(headers) - len(values))
        row_values = dict(zip(headers, padded_values))
        source_smiles = clean_text(padded_values[smiles_index])
        canonical, mol, error = parse_smiles(source_smiles)
        parsed_rows.append(
            MoleculeRow(
                excel_row=output_index + 1,
                source_smiles=source_smiles,
                canonical_smiles=canonical,
                mol=mol,
                error=error,
                values=row_values,
            )
        )

    return headers, parsed_rows


def write_tsv(
    path: Path,
    headers: Sequence[str],
    rows: Iterable[MoleculeRow],
) -> tuple[int, int]:
    """Write canonical SMILES first, followed by all non-SMILES source columns."""
    path.parent.mkdir(parents=True, exist_ok=True)
    source_smiles_columns = {header for header in headers if "smiles" in header.lower()}
    metadata_columns = [header for header in headers if header not in source_smiles_columns]
    output_columns = ["smiles", *metadata_columns, "smiles_valid", "smiles_error", "excel_row"]

    total = 0
    invalid = 0
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=output_columns,
            delimiter="\t",
            lineterminator="\n",
            extrasaction="ignore",
        )
        writer.writeheader()
        for row in rows:
            total += 1
            if row.error:
                invalid += 1
            output = {
                "smiles": row.canonical_smiles or row.source_smiles,
                "smiles_valid": "1" if row.mol is not None else "0",
                "smiles_error": row.error,
                "excel_row": str(row.excel_row),
            }
            output.update({column: clean_text(row.values.get(column)) for column in metadata_columns})
            writer.writerow(output)

    return total, invalid


def molecule_png(mol: Chem.Mol, width: int = 600, height: int = 390) -> bytes:
    depicted = Chem.Mol(mol)
    try:
        AllChem.Compute2DCoords(depicted)
    except Exception:
        pass

    drawer = rdMolDraw2D.MolDraw2DCairo(width, height)
    drawer.drawOptions().padding = 0.08
    try:
        rdMolDraw2D.PrepareAndDrawMolecule(drawer, depicted, kekulize=True)
    except Exception:
        rdMolDraw2D.PrepareAndDrawMolecule(drawer, depicted, kekulize=False)
    drawer.FinishDrawing()
    return drawer.GetDrawingText()


def split_long_token(
    token: str,
    font_name: str,
    font_size: float,
    max_width: float,
) -> list[str]:
    if stringWidth(token, font_name, font_size) <= max_width:
        return [token]

    chunks: list[str] = []
    current = ""
    for character in token:
        trial = current + character
        if current and stringWidth(trial, font_name, font_size) > max_width:
            chunks.append(current)
            current = character
        else:
            current = trial
    if current:
        chunks.append(current)
    return chunks


def wrap_text(
    text: str,
    font_name: str,
    font_size: float,
    max_width: float,
) -> list[str]:
    words: list[str] = []
    for token in text.split():
        words.extend(split_long_token(token, font_name, font_size, max_width))
    if not words:
        return [""]

    lines: list[str] = []
    current = words[0]
    for word in words[1:]:
        trial = f"{current} {word}"
        if stringWidth(trial, font_name, font_size) <= max_width:
            current = trial
        else:
            lines.append(current)
            current = word
    lines.append(current)
    return lines


def row_title(row: MoleculeRow, config: SheetConfig) -> str:
    parts = [
        clean_text(row.values.get(column))
        for column in config.title_columns
        if clean_text(row.values.get(column))
    ]
    return " | ".join(parts) if parts else f"Excel row {row.excel_row}"


def pdf_feature_columns(headers: Sequence[str], config: SheetConfig) -> list[str]:
    smiles_columns = {header for header in headers if "smiles" in header.lower()}
    candidates = list(config.pdf_columns) if config.pdf_columns else list(headers)
    return [column for column in candidates if column in headers and column not in smiles_columns]


def write_pdf(
    path: Path,
    workbook_path: Path,
    headers: Sequence[str],
    config: SheetConfig,
    rows: Iterable[MoleculeRow],
    per_page: int = 6,
) -> tuple[int, int]:
    """Write structures plus table metadata; no SMILES text is included."""
    path.parent.mkdir(parents=True, exist_ok=True)
    document = canvas.Canvas(str(path), pagesize=A4, pageCompression=1)
    page_width, page_height = A4
    margin_x = 10 * mm
    margin_bottom = 10 * mm
    header_height = 13 * mm
    footer_height = 6 * mm
    content_top = page_height - header_height
    content_bottom = margin_bottom + footer_height
    row_height = (content_top - content_bottom) / per_page
    image_width = 66 * mm
    image_height = min(42 * mm, row_height - 3 * mm)
    text_x = margin_x + image_width + 4 * mm
    text_width = page_width - margin_x - text_x
    font_size = 7.2
    line_height = 8.5
    max_lines = max(3, int((row_height - 8 * mm) / line_height))
    feature_columns = pdf_feature_columns(headers, config)

    page_number = 0
    page_open = False
    total = 0
    invalid = 0

    def begin_page() -> None:
        nonlocal page_number, page_open
        page_number += 1
        page_open = True
        document.setFont("Helvetica-Bold", 11)
        document.drawString(margin_x, page_height - 8 * mm, f"{config.name} structure report")
        document.setFont("Helvetica", 7)
        document.drawRightString(
            page_width - margin_x,
            page_height - 8 * mm,
            workbook_path.name,
        )

    def finish_page() -> None:
        nonlocal page_open
        document.setFont("Helvetica", 7)
        document.drawString(margin_x, margin_bottom, "SMILES omitted from PDF; see companion TSV.")
        document.drawRightString(page_width - margin_x, margin_bottom, f"Page {page_number}")
        document.showPage()
        page_open = False

    for row in rows:
        entry_index = total % per_page
        if entry_index == 0:
            if page_open:
                finish_page()
            begin_page()

        total += 1
        if row.error:
            invalid += 1

        y_top = content_top - entry_index * row_height
        image_y = y_top - image_height - 1 * mm
        if row.mol is not None:
            png = molecule_png(row.mol)
            document.drawImage(
                ImageReader(BytesIO(png)),
                margin_x,
                image_y,
                width=image_width,
                height=image_height,
                preserveAspectRatio=True,
                anchor="c",
                mask="auto",
            )
        else:
            document.setFont("Helvetica-Oblique", 9)
            document.drawCentredString(
                margin_x + image_width / 2,
                image_y + image_height / 2,
                f"Invalid structure: {row.error}",
            )

        title = row_title(row, config)
        document.setFont("Helvetica-Bold", 8.5)
        title_lines = wrap_text(title, "Helvetica-Bold", 8.5, text_width)
        text_y = y_top - 3 * mm
        for line in title_lines[:2]:
            document.drawString(text_x, text_y, line)
            text_y -= 10

        features = "   ".join(
            f"{column}={clean_text(row.values.get(column))}"
            for column in feature_columns
            if column not in config.title_columns and clean_text(row.values.get(column))
        )
        document.setFont("Helvetica", font_size)
        feature_lines = wrap_text(features, "Helvetica", font_size, text_width)
        lines_available = max(1, max_lines - min(2, len(title_lines)))
        visible_lines = feature_lines[:lines_available]
        if len(feature_lines) > lines_available and visible_lines:
            visible_lines[-1] = f"{visible_lines[-1][:-3]}..." if len(visible_lines[-1]) > 3 else "..."
        for line in visible_lines:
            document.drawString(text_x, text_y, line)
            text_y -= line_height

        separator_y = y_top - row_height + 1 * mm
        document.setStrokeColorRGB(0.75, 0.75, 0.75)
        document.setLineWidth(0.35)
        document.line(margin_x, separator_y, page_width - margin_x, separator_y)

    if page_open:
        finish_page()
    document.save()
    return total, invalid


def convert_sheet(
    workbook: Any,
    workbook_path: Path,
    output_dir: Path,
    config: SheetConfig,
    *,
    limit: int | None,
    skip_pdf: bool,
) -> tuple[int, int]:
    stem = safe_stem(config.name)
    tsv_path = output_dir / f"{stem}.tsv"
    pdf_path = output_dir / f"{stem}_structures.pdf"

    headers, rows = read_sheet_rows(workbook, config, limit=limit)
    total, invalid = write_tsv(tsv_path, headers, rows)
    print(f"[OK] {config.name}: wrote {total} rows to {tsv_path}")

    if not skip_pdf:
        _, pdf_invalid = write_pdf(
            pdf_path,
            workbook_path,
            headers,
            config,
            rows,
        )
        if pdf_invalid != invalid:
            raise RuntimeError(f"Validation counts differ while writing {config.name}")
        print(f"[OK] {config.name}: wrote {total} structures to {pdf_path}")

    if invalid:
        print(f"[WARN] {config.name}: {invalid} invalid or empty SMILES", file=sys.stderr)
    return total, invalid


def parse_args() -> argparse.Namespace:
    script_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(
        description=(
            "Read table S1/S3/S4 from an Excel workbook and create structure-only "
            "PDF reports plus canonical-SMILES TSV files."
        )
    )
    parser.add_argument(
        "workbook",
        nargs="?",
        type=Path,
        default=script_dir / DEFAULT_WORKBOOK,
        help=f"Input XLSX workbook (default: {DEFAULT_WORKBOOK} next to this script)",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=script_dir / "converted",
        help="Output directory (default: 10062026/converted)",
    )
    parser.add_argument(
        "--sheets",
        nargs="+",
        choices=[config.name for config in SHEET_CONFIGS],
        default=[config.name for config in SHEET_CONFIGS],
        help="Worksheets to convert (default: all configured sheets)",
    )
    parser.add_argument(
        "--limit",
        type=int,
        help="Process only the first N data rows per sheet (useful for testing)",
    )
    parser.add_argument(
        "--skip-pdf",
        action="store_true",
        help="Write TSV files only",
    )
    args = parser.parse_args()
    if args.limit is not None and args.limit < 1:
        parser.error("--limit must be at least 1")
    return args


def main() -> int:
    args = parse_args()
    workbook_path = args.workbook.expanduser().resolve()
    output_dir = args.output_dir.expanduser().resolve()
    if not workbook_path.is_file():
        print(f"Workbook not found: {workbook_path}", file=sys.stderr)
        return 2

    selected = {name for name in args.sheets}
    totals = []
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        for config in SHEET_CONFIGS:
            if config.name in selected:
                totals.append(
                    convert_sheet(
                        workbook,
                        workbook_path,
                        output_dir,
                        config,
                        limit=args.limit,
                        skip_pdf=args.skip_pdf,
                    )
                )
    finally:
        workbook.close()

    row_count = sum(total for total, _ in totals)
    invalid_count = sum(invalid for _, invalid in totals)
    print(
        f"[DONE] Processed {row_count} rows across {len(totals)} sheets; "
        f"{invalid_count} invalid or empty SMILES."
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
